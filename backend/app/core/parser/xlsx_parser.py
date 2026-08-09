"""XLSX 解析器。"""
from openpyxl import load_workbook

from app.core.parser.models import ParsedElement


async def parse(path: str) -> list[ParsedElement]:
    """解析 xlsx：每个工作表转为一个 table 元素（HTML），section_path 为表名。"""
    elements: list[ParsedElement] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = [[("" if c is None else str(c)) for c in row] for row in ws.iter_rows(values_only=True)]
        if not rows:
            continue
        body = "".join("<tr>" + "".join(f"<td>{c}</td>" for c in r) + "</tr>" for r in rows)
        elements.append(ParsedElement("table", f"<table>{body}</table>", 1, section_path=ws.title))
    wb.close()
    return elements
